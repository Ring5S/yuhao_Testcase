import openpyxl


def main():
    someone = 'you'
    while someone == 'you':
        print("I LOVE {someone}".format(someone=someone))


# 实例化
wb = openpyxl.Workbook()
# 激活 worksheet
ws = wb.active
"""打开已有表格文件"""
wb2 = openpyxl.load_workbook(r'E:/Test_1/test1.xlsx')
ws1 = wb.create_sheet("sheet2")
print(wb.sheetnames)


